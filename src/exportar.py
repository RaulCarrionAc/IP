"""
Exportación del reporte de IP a Excel.

Genera un .xlsx con el detalle de resultados en la hoja "Reporte" y,
debajo de la tabla, una fila con el promedio de IP destacado en negrita.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font


def exportar_reporte_excel(
    df_resultados: pd.DataFrame,
    salida: Path,
    empresa: str,
    mes: str,
    anio: str | int,
) -> Path:
    """
    Exporta df_resultados a un Excel y agrega el promedio de IP debajo
    de la tabla.

    Parámetros
    ----------
    df_resultados : DataFrame con la columna "IP" (resultado de calcular_resultados).
    salida : carpeta donde se guardará el archivo.
    empresa, mes, anio : se usan para componer el nombre del archivo.

    Retorna
    -------
    Path del archivo Excel generado.
    """
    # 1. Calcular el promedio general de la variable deseada
    ip_promedio = df_resultados["IP"].mean()

    # 2. Definir el nombre del archivo de salida
    salida = Path(salida)
    salida.mkdir(parents=True, exist_ok=True)
    nombre_archivo_excel = salida / f"reporte_IP_{empresa}_{mes}{anio}.xlsx"

    # 3. Exportar usando el motor 'openpyxl' para poder modificar celdas
    #    después de escribir el DataFrame
    with pd.ExcelWriter(nombre_archivo_excel, engine="openpyxl") as writer:
        # Guardamos el DataFrame normal en la pestaña 'Reporte'
        df_resultados.to_excel(writer, sheet_name="Reporte", index=False)

        # Obtenemos la hoja de trabajo activa para agregar las filas adicionales
        worksheet = writer.sheets["Reporte"]

        # Calculamos la fila exacta: Total de datos + 1 fila de encabezado
        # + 2 filas de separación (salto)
        fila_salto = len(df_resultados) + 3

        # Escribimos el texto y el valor del promedio en esa fila
        worksheet.cell(row=fila_salto, column=1, value="Promedio IP:")
        worksheet.cell(row=fila_salto, column=2, value=ip_promedio)

        # Poner en negrita la celda del resultado para que destaque
        worksheet.cell(row=fila_salto, column=1).font = Font(bold=True)
        worksheet.cell(row=fila_salto, column=2).font = Font(bold=True)

    print(f"Reporte exportado exitosamente a {nombre_archivo_excel}")

    return nombre_archivo_excel
